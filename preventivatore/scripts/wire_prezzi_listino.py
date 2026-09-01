# -*- coding: utf-8 -*-
"""
rev.28 — collega LISTINO!G (Prezzo unitario €) alle schede.

Il file oggi ha la formula del prezzo solo per la bulloneria: per carpenteria,
pezzi speciali ed elettromeccanica la colonna G e' vuota, quindi tutta la
catena DISTINTA!Q -> DISTINTA!R -> DISTINTA!R346 resta ferma.

Questo script parte dalla revisione piu' alta presente nel repo, scrive in
LISTINO!G una cascata di lookup sulle quattro schede e salva una NUOVA
revisione (rev.NN+1). La revisione di partenza non viene toccata.

  Excel COM non e' disponibile: si usa openpyxl (le formule le calcola Excel
  all'apertura del file).

Uso:
    python preventivatore/scripts/wire_prezzi_listino.py
    python preventivatore/scripts/wire_prezzi_listino.py --src "../LISTINO ... rev.27 ....xlsx" --out "../LISTINO ... rev.28 ....xlsx"
"""
from __future__ import annotations
import argparse
import glob
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl non installato:  pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Ancore dei blocchi 'Elenco' nelle schede (verificate su rev.27) -------------
STRUTTURA_RANGE = "'SCHEDA STRUTTURA'!$A$17:$G$147"        # A codice ... F costo acq, G costo zinc
STRUTTURA_KEYS = "'SCHEDA STRUTTURA'!$A$17:$A$147"
PEZZI_RANGE = "'SCHEDA PEZZI SPECIALI'!$A$38:$F$59"        # A codice ... E costo acq, F costo zinc
PEZZI_KEYS = "'SCHEDA PEZZI SPECIALI'!$A$38:$A$59"
ELETTRO_RANGE = "'SCHEDA ELETTROMECCANICA'!$A$5:$D$14"     # A codice ... D costo (EUR/pz)
ELETTRO_KEYS = "'SCHEDA ELETTROMECCANICA'!$A$5:$A$14"
BULL_RANGE = "'SCHEDA BULLONERIA'!$A$6:$D$36"              # A codice ... D costo (EUR/pz)
BULL_KEYS = "'SCHEDA BULLONERIA'!$A$6:$A$36"


def price_formula(row: int) -> str:
    b = f"$B{row}"
    return (
        f'=IF({b}="","",'
        f'IF(ISNUMBER(MATCH({b},{STRUTTURA_KEYS},0)),'
        f'IFERROR(VLOOKUP({b},{STRUTTURA_RANGE},6,FALSE)*1,0)+IFERROR(VLOOKUP({b},{STRUTTURA_RANGE},7,FALSE)*1,0),'
        f'IF(ISNUMBER(MATCH({b},{PEZZI_KEYS},0)),'
        f'IFERROR(VLOOKUP({b},{PEZZI_RANGE},5,FALSE)*1,0)+IFERROR(VLOOKUP({b},{PEZZI_RANGE},6,FALSE)*1,0),'
        f'IF(ISNUMBER(MATCH({b},{ELETTRO_KEYS},0)),'
        f'IFERROR(VLOOKUP({b},{ELETTRO_RANGE},4,FALSE)*1,0),'
        f'IF(ISNUMBER(MATCH({b},{BULL_KEYS},0)),'
        f'IFERROR(VLOOKUP({b},{BULL_RANGE},4,FALSE)*1,0),'
        f'"")))))'
    )


def find_latest_xlsx() -> str:
    cands = glob.glob(os.path.join(REPO, "LISTINO componenti tracker TTS 1303 - rev.*.xlsx"))
    cands = [c for c in cands if "vecchio" not in c.lower()]
    if not cands:
        sys.exit("Nessun file 'LISTINO ... rev.N.xlsx' nella cartella del repo.")
    return max(cands, key=lambda p: int(re.search(r"rev\.(\d+)", os.path.basename(p)).group(1)))


def bump_name(src: str) -> str:
    base = os.path.basename(src)
    m = re.search(r"rev\.(\d+)", base)
    n = int(m.group(1)) + 1
    return os.path.join(
        REPO,
        f"LISTINO componenti tracker TTS 1303 - rev.{n} (prezzo unitario listino collegato alle schede).xlsx",
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    src = args.src or find_latest_xlsx()
    out = args.out or bump_name(src)
    if os.path.abspath(src) == os.path.abspath(out):
        sys.exit("out coincide con src: scegli un nome diverso, la revisione di partenza non va toccata.")
    print(f"Sorgente : {os.path.basename(src)}")
    print(f"Output   : {os.path.basename(out)}")

    wb = load_workbook(src)
    ws = wb["LISTINO"]

    # header 'Prezzo unitario €' atteso in G4
    if str(ws["G4"].value).strip().lower() not in ("prezzo unitario €", "prezzo unitario"):
        sys.exit(f"G4 = {ws['G4'].value!r}: layout LISTINO inatteso, interrompo per sicurezza.")

    written = skipped = 0
    for r in range(5, ws.max_row + 1):
        code = ws.cell(r, 2).value  # col B
        if code in (None, ""):
            skipped += 1
            continue
        c = ws.cell(r, 7)  # col G
        c.value = price_formula(r)
        if not c.number_format or c.number_format == "General":
            c.number_format = "#,##0.00"
        written += 1

    wb.save(out)
    print(f"\nLISTINO!G : {written} formule scritte, {skipped} righe senza codice saltate.")
    print("Catena che si attiva al ricalcolo in Excel:")
    print("   LISTINO!G  ->  DISTINTA COMPLETA!Q  ->  DISTINTA COMPLETA!R  ->  R346 (TOTALE)")
    print("\nControlli rapidi attesi (dopo 'Calcola ora' in Excel):")
    for code, val in (("TTS.AR.001.7200", "106,20"), ("TTS.PC.001.IPEA140.39", "66,42"),
                      ("TTS.PF.001.IPEA140.36", "61,31"), ("TTS.PM.002.IPE140", "8,33")):
        print(f"   {code:<26} G = {val}")
    print("   Restano vuoti: 31 bulloneria (manca 'Costo €/pz' in SCHEDA BULLONERIA) + 8 elettrici/sensori + 2 varianti.")


if __name__ == "__main__":
    main()
