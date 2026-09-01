# -*- coding: utf-8 -*-
"""
ETL — LISTINO/DISTINTA workbook  ->  seed data per il preventivatore.

Legge l'ultima revisione del file Excel (rev.27) e produce:
  preventivatore/src/data/components.json      (anagrafica componenti + costo calcolato)
  preventivatore/src/data/bom.json             (righe di distinta, con quantita' per configurazione)
  preventivatore/src/data/tracker_configs.json (le 6 taglie di tracker)
  preventivatore/supabase/seed.sql             (INSERT per Supabase)
  preventivatore/src/data/_meta.json           (statistiche + copertura prezzi)

Excel COM non e' disponibile su questa postazione: si usa openpyxl in sola lettura.

Uso:
    python preventivatore/scripts/etl.py
    python preventivatore/scripts/etl.py --xlsx "../LISTINO ... rev.28 ....xlsx"
"""
from __future__ import annotations
import argparse
import glob
import json
import os
import re
import sys
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl non installato:  pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))      # preventivatore/
REPO = os.path.dirname(ROOT)
DATA = os.path.join(ROOT, "src", "data")
SUPA = os.path.join(ROOT, "supabase")

MODULE_SIZES = [12, 18, 20, 22, 36, 40]
CONFIG_COLS = {12: 7, 18: 8, 20: 9, 22: 10, 36: 11, 40: 12}   # 0-based col index in DISTINTA COMPLETA

# --- Perimetro di preventivo (default; sovrascrivibile nella app) --------------
EXCLUDED_ASSEMBLIES = {
    "kit quadri area",            # quadro area  -> fuori perimetro
    "kit quadri centralina meteo",  # quadri generali / centralina meteo
    "kit quadri inverter",
    "sensori meteo",             # sensoristica meteo
    "kit quadro motore",         # sostituito da voce forfettaria 170 EUR/tracker
}
EXCLUDED_CATEGORIES = {
    "Quadristica ed elettrico",
    "Moduli fotovoltaici",       # pannelli forniti dal cliente (default OFF)
}
QUADRO_MOTORE_EUR = 170.0        # una voce per ogni tracker
BULLONERIA_EUR_KG = 3.0         # stima di ripiego per la minuteria priva di prezzo

# price_source -> affidabilita' del prezzo (guida la "puntualita'" del preventivo)
CONFIDENCE = {
    "confermato": "alta",             # inserito/validato a mano in app
    "listino": "alta",                # prezzo gia' compilato nel foglio LISTINO
    "scheda_commerciale": "alta",     # articolo commerciale con prezzo a catalogo
    "scheda_struttura": "media",      # calcolato: peso x EUR/kg (acquisto + zincatura)
    "scheda_pezzi_speciali": "media", # calcolato: peso x EUR/kg
    "stima_peso": "bassa",            # bulloneria: peso x EUR/kg forfettario
    "mancante": "nulla",
    "senza_codice": "nulla",
}


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def find_latest_xlsx() -> str:
    cands = glob.glob(os.path.join(REPO, "LISTINO componenti tracker TTS 1303 - rev.*.xlsx"))
    if not cands:
        sys.exit("Nessun file 'LISTINO ... rev.N.xlsx' trovato nella cartella del repo.")

    def rev(p):
        m = re.search(r"rev\.(\d+)", os.path.basename(p))
        return int(m.group(1)) if m else -1

    return max(cands, key=rev)


# --------------------------------------------------------------------------- #
#  Estrazione costi calcolati dalle "schede"                                  #
# --------------------------------------------------------------------------- #
def read_scheda_struttura(wb):
    """SCHEDA STRUTTURA, blocco 'Elenco': Codice | Descr | Profilo | Lungh | Peso | Costo acq | Costo zinc"""
    ws = wb["SCHEDA STRUTTURA"]
    out = {}
    started = False
    for row in ws.iter_rows(values_only=True):
        c = norm(row[0])
        if c == "Codice":
            started = True
            continue
        if not started or not c:
            continue
        out[c] = {
            "peso_kg": num(row[4]),
            "costo_acquisto": num(row[5]),
            "costo_zincatura": num(row[6]) or 0.0,
        }
    return out


def read_scheda_pezzi(wb):
    """SCHEDA PEZZI SPECIALI, blocco 'Elenco': Codice | Descr | Voce | Peso | Costo acq | Costo zinc"""
    ws = wb["SCHEDA PEZZI SPECIALI"]
    out = {}
    started = False
    for row in ws.iter_rows(values_only=True):
        c = norm(row[0])
        if c == "Codice":
            started = True
            continue
        if not started or not c:
            continue
        out[c] = {
            "peso_kg": num(row[3]),
            "costo_acquisto": num(row[4]),
            "costo_zincatura": num(row[5]) or 0.0,
        }
    return out


def read_scheda_commerciale(wb):
    """SCHEDA BULLONERIA + SCHEDA ELETTROMECCANICA: Codice | Descr | Peso unit | Costo EUR/pz"""
    out = {}
    for name in ("SCHEDA BULLONERIA", "SCHEDA ELETTROMECCANICA"):
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            c = norm(row[0])
            if not c or c == "Codice" or c.startswith("SCHEDA") or c.startswith("BULLONERIA"):
                continue
            out[c] = {"peso_kg": num(row[2]), "prezzo": num(row[3])}
    return out


# --------------------------------------------------------------------------- #
#  Componenti (LISTINO)                                                       #
# --------------------------------------------------------------------------- #
def build_components(wb, struttura, pezzi, commerciale):
    ws = wb["LISTINO"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows) if norm(r[0]) == "N°")
    comps = []
    for r in rows[hdr_i + 1:]:
        code = norm(r[1])
        desc = norm(r[2])
        if not code and not desc:
            continue
        cat = norm(r[3]) or "Altro"
        weight_listino = num(r[5])
        price_listino = num(r[6])

        calc = None
        source = "mancante"
        if code in struttura and struttura[code]["costo_acquisto"] is not None:
            s = struttura[code]
            calc = round(s["costo_acquisto"] + s["costo_zincatura"], 4)
            source = "scheda_struttura"
        elif code in pezzi and pezzi[code]["costo_acquisto"] is not None:
            s = pezzi[code]
            calc = round(s["costo_acquisto"] + s["costo_zincatura"], 4)
            source = "scheda_pezzi_speciali"
        elif code in commerciale and commerciale[code]["prezzo"] is not None:
            calc = commerciale[code]["prezzo"]
            source = "scheda_commerciale"
        elif price_listino is not None:
            calc = price_listino
            source = "listino"
        elif cat == "Bulloneria e minuteria":
            w = weight_listino or (commerciale.get(code) or {}).get("peso_kg")
            if w:
                calc = round(w * BULLONERIA_EUR_KG, 4)
                source = "stima_peso"

        comps.append({
            "code": code or None,
            "description": desc,
            "category": cat,
            "uom": norm(r[4]) or "pz",
            "weight_kg": weight_listino,
            "price_listino": price_listino,
            "computed_cost": calc,
            "price_source": source,     # confermato | scheda_* | listino | stima_peso | mancante
            "price_confidence": CONFIDENCE.get(source, "nulla"),
        })
    return comps


# --------------------------------------------------------------------------- #
#  Distinta (DISTINTA COMPLETA)                                               #
# --------------------------------------------------------------------------- #
def build_bom(wb):
    ws = wb["DISTINTA COMPLETA"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for i, r in enumerate(rows[1:], start=2):
        asm = norm(r[0])
        desc = norm(r[4])
        if not asm and not norm(r[3]) and "TOTALE" in desc.upper():
            continue
        if not asm:
            continue
        qb = num(r[6])
        by_cfg = {str(m): num(r[CONFIG_COLS[m]]) for m in MODULE_SIZES}
        out.append({
            "row": i,
            "assembly": asm,
            "section": norm(r[1]) or None,
            "position": norm(r[2]) or None,
            "code": norm(r[3]) or None,
            "description": desc,
            "category": norm(r[5]) or "Altro",
            "qty_base": qb,
            "qty_by_config": by_cfg,
            "already_loaded": norm(r[13]).upper() == "X",
        })
    return out


# --------------------------------------------------------------------------- #
#  Motore di preventivo (Python) — usato solo per la validazione a schermo    #
# --------------------------------------------------------------------------- #
def price_of(comp):
    return comp["computed_cost"] if comp["computed_cost"] is not None else 0.0


def explode_tracker(bom, comps_by_code, modules):
    asse_section = f"{modules} MODULI 1303"
    lines = []
    for bl in bom:
        if bl["assembly"] in EXCLUDED_ASSEMBLIES:
            continue
        if bl["already_loaded"]:
            continue
        if bl["category"] in EXCLUDED_CATEGORIES:
            continue
        if bl["assembly"] == "ASSI DI ROTAZIONE 1303" and bl["section"] != asse_section:
            continue
        q = bl["qty_by_config"].get(str(modules))
        if q is None:
            q = bl["qty_base"]
        if not q:
            continue
        comp = comps_by_code.get(bl["code"]) if bl["code"] else None
        unit = price_of(comp) if comp else 0.0
        src = comp["price_source"] if comp else "senza_codice"
        lines.append({
            "code": bl["code"], "description": bl["description"],
            "qty": q, "unit_cost": unit, "amount": round(q * unit, 2),
            "price_source": src, "confidence": CONFIDENCE.get(src, "nulla"),
            "priced": unit > 0,
        })
    return lines


def quote(bom, comps, project):
    by_code = {c["code"]: c for c in comps if c["code"]}
    out = {"tracker_types": [], "extras": [], "totals": {}}
    material_total = 0.0
    priced_amount = 0.0
    missing_amount_lines = 0
    total_lines = 0
    conf_amount = {"alta": 0.0, "media": 0.0, "bassa": 0.0, "nulla": 0.0}
    for t in project["tracker_types"]:
        lines = explode_tracker(bom, by_code, t["modules"])
        mat = round(sum(l["amount"] for l in lines), 2)
        priced = round(sum(l["amount"] for l in lines if l["priced"]), 2)
        for l in lines:
            conf_amount[l["confidence"]] = conf_amount.get(l["confidence"], 0.0) + l["amount"] * t["count"]
        n_missing = sum(1 for l in lines if not l["priced"] and l["qty"])
        quadro = QUADRO_MOTORE_EUR * t["count"]
        per_tracker = round(mat, 2)
        subtotal = round(per_tracker * t["count"] + quadro, 2)
        material_total += per_tracker * t["count"]
        priced_amount += priced * t["count"]
        missing_amount_lines += n_missing
        total_lines += len(lines)
        out["tracker_types"].append({
            "modules": t["modules"], "count": t["count"],
            "material_per_tracker": per_tracker,
            "quadro_motore": quadro,
            "subtotal": subtotal,
            "lines_total": len(lines), "lines_missing_price": n_missing,
        })
    quadro_total = QUADRO_MOTORE_EUR * sum(t["count"] for t in project["tracker_types"])
    out["extras"].append({"label": "Quadro motore (forfait)", "unit": QUADRO_MOTORE_EUR,
                          "qty": sum(t["count"] for t in project["tracker_types"]),
                          "amount": round(quadro_total, 2)})
    cost = round(material_total + quadro_total, 2)
    margin = project.get("margin_pct", 0) / 100.0
    sell = round(cost * (1 + margin), 2)
    out["totals"] = {
        "n_tracker": sum(t["count"] for t in project["tracker_types"]),
        "material_cost": round(material_total, 2),
        "quadro_motore_cost": round(quadro_total, 2),
        "cost_total": cost,
        "margin_pct": project.get("margin_pct", 0),
        "sell_total": sell,
        "price_coverage_pct": round(100.0 * priced_amount / material_total, 1) if material_total else 0.0,
        "confidence_amount": {k: round(v, 2) for k, v in conf_amount.items()},
        "confidence_pct": {k: (round(100.0 * v / material_total, 1) if material_total else 0.0)
                           for k, v in conf_amount.items()},
        "bom_lines": total_lines,
        "bom_lines_missing_price": missing_amount_lines,
    }
    return out


# --------------------------------------------------------------------------- #
def sql_escape(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"


def write_seed_sql(comps, bom, configs, path):
    lines = ["-- Generato da scripts/etl.py — non modificare a mano", "begin;",
             "truncate table bom_lines, components, tracker_configs restart identity cascade;", ""]
    for c in comps:
        lines.append(
            "insert into components (code, description, category, uom, weight_kg, price_listino, computed_cost, price_source, price_confidence) values ("
            + ", ".join(sql_escape(c[k]) for k in
                        ("code", "description", "category", "uom", "weight_kg", "price_listino", "computed_cost", "price_source", "price_confidence"))
            + ");")
    lines.append("")
    for b in bom:
        lines.append(
            "insert into bom_lines (src_row, assembly, section, position, code, description, category, qty_base, qty_by_config, already_loaded) values ("
            + ", ".join([sql_escape(b["row"]), sql_escape(b["assembly"]), sql_escape(b["section"]),
                         sql_escape(b["position"]), sql_escape(b["code"]), sql_escape(b["description"]),
                         sql_escape(b["category"]), sql_escape(b["qty_base"]),
                         "'" + json.dumps(b["qty_by_config"]) + "'::jsonb", sql_escape(b["already_loaded"])])
            + ");")
    lines.append("")
    for cfg in configs:
        lines.append(
            "insert into tracker_configs (modules, asse_section, label) values ("
            + ", ".join([sql_escape(cfg["modules"]), sql_escape(cfg["asse_section"]), sql_escape(cfg["label"])])
            + ");")
    lines.append("commit;")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None, help="percorso del file Excel (default: rev. piu' alta nel repo)")
    args = ap.parse_args()

    xlsx = args.xlsx or find_latest_xlsx()
    print(f"Sorgente : {os.path.basename(xlsx)}")
    wb = load_workbook(xlsx, data_only=True)

    struttura = read_scheda_struttura(wb)
    pezzi = read_scheda_pezzi(wb)
    commerciale = read_scheda_commerciale(wb)
    comps = build_components(wb, struttura, pezzi, commerciale)
    bom = build_bom(wb)
    configs = [{"modules": m, "asse_section": f"{m} MODULI 1303", "label": f"Tracker {m} moduli"}
               for m in MODULE_SIZES]

    os.makedirs(DATA, exist_ok=True)
    os.makedirs(SUPA, exist_ok=True)
    with open(os.path.join(DATA, "components.json"), "w", encoding="utf-8") as f:
        json.dump(comps, f, ensure_ascii=False, indent=1)
    with open(os.path.join(DATA, "bom.json"), "w", encoding="utf-8") as f:
        json.dump(bom, f, ensure_ascii=False, indent=1)
    with open(os.path.join(DATA, "tracker_configs.json"), "w", encoding="utf-8") as f:
        json.dump(configs, f, ensure_ascii=False, indent=1)
    write_seed_sql(comps, bom, configs, os.path.join(SUPA, "seed.sql"))

    # ---- copertura prezzi -------------------------------------------------- #
    by_src = {}
    for c in comps:
        by_src[c["price_source"]] = by_src.get(c["price_source"], 0) + 1
    priced = sum(1 for c in comps if c["computed_cost"] not in (None, 0))

    # ---- progetto di esempio (dalla richiesta) --------------------------- #
    example = {
        "name": "Progetto di esempio",
        "margin_pct": 0,
        "tracker_types": [
            {"modules": 12, "count": 4},
            {"modules": 18, "count": 1},
            {"modules": 20, "count": 20},
            {"modules": 22, "count": 5},
            {"modules": 36, "count": 34},
            {"modules": 40, "count": 5},
        ],
    }
    q = quote(bom, comps, example)

    meta = {
        "generated": date.today().isoformat(),
        "source_file": os.path.basename(xlsx),
        "counts": {"components": len(comps), "bom_lines": len(bom), "priced_components": priced},
        "price_source_breakdown": by_src,
        "excluded_assemblies": sorted(EXCLUDED_ASSEMBLIES),
        "excluded_categories": sorted(EXCLUDED_CATEGORIES),
        "quadro_motore_eur": QUADRO_MOTORE_EUR,
        "example_project": example,
        "example_quote": q,
    }
    with open(os.path.join(DATA, "_meta.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=1)

    # ---- dati per la demo HTML (script classico, apribile con doppio clic) - #
    demo_blob = {
        "generated": meta["generated"],
        "source_file": meta["source_file"],
        "components": comps,
        "bom": bom,
        "configs": [{"modules": m, "asseSection": f"{m} MODULI 1303"} for m in MODULE_SIZES],
        "excludedAssemblies": sorted(EXCLUDED_ASSEMBLIES),
        "excludedCategories": sorted(EXCLUDED_CATEGORIES),
        "quadroMotoreEur": QUADRO_MOTORE_EUR,
        "exampleProject": example,
    }
    with open(os.path.join(REPO, "preventivo-demo-data.js"), "w", encoding="utf-8") as f:
        f.write("/* Generato da preventivatore/scripts/etl.py — non modificare a mano */\n")
        f.write("window.__DEMO__ = ")
        json.dump(demo_blob, f, ensure_ascii=False)
        f.write(";\n")

    # ---- report a schermo ----------------------------------------------- #
    print(f"\nComponenti : {len(comps)}   (con costo: {priced}  /  senza: {len(comps)-priced})")
    for k, v in sorted(by_src.items(), key=lambda x: -x[1]):
        print(f"    {k:24s} {v}")
    print(f"Righe distinta : {len(bom)}")
    print("\n--- PREVENTIVO DI ESEMPIO --------------------------------------")
    for t in q["tracker_types"]:
        print(f"  {t['count']:>3} x tracker {t['modules']:>2} mod  "
              f"materiale/cad {t['material_per_tracker']:>10.2f}  "
              f"+ quadro {t['quadro_motore']:>8.2f}  = {t['subtotal']:>12.2f}   "
              f"(righe {t['lines_total']}, senza prezzo {t['lines_missing_price']})")
    tot = q["totals"]
    print("  " + "-" * 60)
    print(f"  n. tracker ................ {tot['n_tracker']}")
    print(f"  materiale ................. {tot['material_cost']:>14.2f} EUR")
    print(f"  quadri motore (forfait) .. {tot['quadro_motore_cost']:>14.2f} EUR")
    print(f"  COSTO TOTALE ............. {tot['cost_total']:>14.2f} EUR")
    cp = tot["confidence_pct"]
    print(f"  affidabilita' prezzi ..... alta {cp['alta']}%  media {cp['media']}%  "
          f"bassa {cp['bassa']}%  nulla {cp['nulla']}%")
    print(f"  righe senza prezzo ....... {tot['bom_lines_missing_price']}/{tot['bom_lines']}")
    print("\nScritti:")
    for p in ("preventivatore/src/data/components.json", "preventivatore/src/data/bom.json",
              "preventivatore/src/data/tracker_configs.json", "preventivatore/src/data/_meta.json",
              "preventivatore/supabase/seed.sql", "preventivo-demo-data.js"):
        print("   " + p)


if __name__ == "__main__":
    main()
