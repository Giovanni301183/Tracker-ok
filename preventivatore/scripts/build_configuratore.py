# -*- coding: utf-8 -*-
"""
rev.29 — aggiunge CONFIGURATORE e DISTINTA PROGETTO.

Parte da rev.28 (LISTINO!G gia' collegato alle schede) e aggiunge:

  CONFIGURATORE     scheda iniziale: struttura Agri/Basso, trave, spessori,
                    infissione+luce (promemoria), lunghezze pilastro/fondazione
                    scelte dagli elenchi esistenti, tabella "N. tracker" per le
                    6 taglie, riepilogo costi (materiale + 170 EUR/tracker quadro
                    motore + margine).

  DISTINTA PROGETTO una riga per ogni riga di DISTINTA COMPLETA, con:
                    - Q.tà progetto = SUM( N.tracker(taglia) x q.tà(taglia) )
                      solo per le taglie selezionate (N. tracker > 0)
                    - perimetro: fuori i kit quadri/sensori/quadro motore
                    - Agri  -> fondazione + piastre di giunzione attive
                      Basso -> niente fondazione ne' giunzione; pilastro alla
                      lunghezza scelta
                    - spessore asse 3,0/2,5, omega 2,0/1,8, trave IPE/IPEA:
                      cambiano il "Codice progetto" su cui si legge il prezzo
                    - Prezzo da LISTINO!G ; Importo = Q.tà x Prezzo
                    - colonna "In preventivo" per il filtro

Le quantita' per taglia sono pre-calcolate qui (valori, non formule): le uniche
celle "vive" nella DISTINTA PROGETTO sono i riferimenti al CONFIGURATORE.

  Excel COM non e' disponibile: openpyxl scrive, Excel calcola all'apertura.

Uso:  python preventivatore/scripts/build_configuratore.py
"""
from __future__ import annotations
import glob
import os
import re
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl non installato:  pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

SIZES = [12, 18, 20, 22, 36, 40]
CFG_COL = {12: 8, 18: 9, 20: 10, 22: 11, 36: 12, 40: 13}   # H..M in DISTINTA COMPLETA
GIA_COL = 14
EXCLUDED_ASSEMBLIES = {
    "kit quadro motore", "kit quadri centralina meteo",
    "kit quadri area", "kit quadri inverter", "sensori meteo",
}
QUADRO_MOTORE_EUR = 170

# lunghezze disponibili (suffissi di codice) --------------------------------- #
LEN_FOND = [f"{n}" for n in range(26, 37)]                 # TTS.PF.001.*.26 .. .36
LEN_CUSC = ["23"] + [f"{n}" for n in range(28, 46)]        # TTS.PC.001.*.23,28..45
LEN_MOT = ["22"] + [f"{n}" for n in range(27, 45)]         # TTS.PM.001.*.22,27..44


def latest(patt: str) -> str:
    c = [x for x in glob.glob(os.path.join(REPO, patt)) if "vecchio" not in x.lower()]
    if not c:
        sys.exit(f"nessun file per {patt!r}")
    return max(c, key=lambda p: int(re.search(r"rev\.(\d+)", os.path.basename(p)).group(1)))


# --------------------------------------------------------------------------- #
#  Tag di riga (con assert: se il file cambia, lo script si ferma)           #
# --------------------------------------------------------------------------- #
CARRIERS = {                       # riga -> codice atteso (assert layout)
    2: "TTS.PF.001.IPE140.30",      # fondazione pilastro motore
    24: "TTS.PM.001.IPE140.22",     # pilastro motore
    83: "TTS.PF.001.IPE140.30",     # fondazione pilastro cuscinetto
    84: "TTS.PC.001.IPE140.23",     # pilastro cuscinetto
    226: "TTS.AR.3014",             # omega supporto pannello
}


def build_tags(ws):
    """ritorna dict row -> (tag, codice_D) sulle righe 2..344 di DISTINTA COMPLETA"""
    def code(r):
        return (ws.cell(r, 4).value or "")

    for r, exp in CARRIERS.items():
        if code(r) != exp:
            sys.exit(f"riga {r}: atteso {exp!r}, trovato {code(r)!r} — layout DISTINTA cambiato, mi fermo.")

    tags = {}
    for r in range(2, 345):
        a = (ws.cell(r, 1).value or "").strip()
        d = code(r)
        gia = str(ws.cell(r, GIA_COL).value or "").strip().upper() == "X"

        if not a:
            tags[r] = ("SKIP", d)
        elif gia or a in EXCLUDED_ASSEMBLIES:
            tags[r] = ("EXCL", d)
        elif a == "PILASTRO MOTORE":
            if r == 2:
                tags[r] = ("FOND_MOT", d)
            elif 3 <= r <= 23:
                tags[r] = ("SKIP", d)                       # varianti fondazione
            elif r == 24:
                tags[r] = ("PIL_MOT", d)
            elif 25 <= r <= 61:
                tags[r] = ("SKIP", d)                       # varianti pilastro motore
            elif 62 <= r <= 72:
                tags[r] = ("GIUNZ", d)                      # piastre giunzione + bulloni
            else:
                tags[r] = ("", d)                           # 73..82 supporti/motoriduttore
        elif a == "PILASTRO CUSCINETTO":
            if r == 83:
                tags[r] = ("FOND_CUSC", d)
            elif r == 84:
                tags[r] = ("PIL_CUSC", d)
            elif 85 <= r <= 121:
                tags[r] = ("SKIP", d)                       # varianti pilastro cuscinetto
            elif 122 <= r <= 132:
                tags[r] = ("GIUNZ", d)
            else:
                tags[r] = ("", d)                           # 133..143 supporti + cuscinetto
        elif a == "ASSI DI ROTAZIONE 1303":
            if d.endswith(".25"):
                tags[r] = ("SKIP", d)                       # variante 2,5 mm: gestita dal carrier
            elif re.match(r"TTS\.AR\.00[1-4]\.", d):
                tags[r] = ("ASSE_PROF", d)
            else:
                tags[r] = ("", d)                           # TTS.AR.2011, bulloneria: sezione-gated
        elif a == "KIT OMEGA+pannelli":
            if 231 <= r <= 235:
                tags[r] = ("PANNELLI", d)
            elif 236 <= r <= 240:
                tags[r] = ("SKIP", d)                       # pannellino optional
            elif d == "TTS.AR.3014":
                tags[r] = ("OMEGA_SUP", d)
            elif d == "TTS.AR.3014.18":
                tags[r] = ("SKIP", d)
            else:
                tags[r] = ("", d)
        else:
            tags[r] = ("", d)                               # KIT FINE CORSA, KIT INCLINOMETRO
    return tags


def qty_by_size(ws, r):
    """q.tà per taglia della riga r: colonna 'Q.tà NN mod' se numerica, altrimenti base G;
       0 se la riga e' di ASSI e la Sezione non e' quella della taglia."""
    a = (ws.cell(r, 1).value or "").strip()
    b = (ws.cell(r, 2).value or "")
    base = ws.cell(r, 7).value
    base = base if isinstance(base, (int, float)) else 0
    out = {}
    for s in SIZES:
        v = ws.cell(r, CFG_COL[s]).value
        q = v if isinstance(v, (int, float)) else base
        if a == "ASSI DI ROTAZIONE 1303" and b != f"{s} MODULI 1303":
            q = 0
        out[s] = q
    return out


# --------------------------------------------------------------------------- #
def main():
    src = latest("LISTINO componenti tracker TTS 1303 - rev.*.xlsx")
    n = int(re.search(r"rev\.(\d+)", os.path.basename(src)).group(1)) + 1
    out = os.path.join(REPO, f"LISTINO componenti tracker TTS 1303 - rev.{n} (configuratore + distinta progetto).xlsx")
    print(f"Sorgente : {os.path.basename(src)}")
    print(f"Output   : {os.path.basename(out)}")

    wb = load_workbook(src)
    if "CONFIGURATORE" in wb.sheetnames or "DISTINTA PROGETTO" in wb.sheetnames:
        sys.exit("il file contiene gia' CONFIGURATORE / DISTINTA PROGETTO.")
    dc = wb["DISTINTA COMPLETA"]
    tags = build_tags(dc)

    # ------------------------------------------------------------------ CONFIGURATORE
    cfg = wb.create_sheet("CONFIGURATORE", 0)
    H = Font(bold=True)
    cfg["A1"] = "CONFIGURATORE PREVENTIVO"
    cfg["A1"].font = Font(bold=True, size=14)

    rows_par = [
        ("A3", "Struttura", "B3", "Agrivoltaico", "Cfg_Struttura", '"Agrivoltaico,Basso"'),
        ("A4", "Trave", "B4", "IPE 140", "Cfg_Trave", '"IPE 140,IPEA 140"'),
        ("A5", "Spessore asse", "B5", "3,0 mm", "Cfg_SpAsse", '"3,0 mm,2,5 mm"'),
        ("A6", "Spessore omega", "B6", "2,0 mm", "Cfg_SpOmega", '"2,0 mm,1,8 mm"'),
        ("A7", "Altezza infissione (m) - promemoria", "B7", 1.5, "Cfg_Infissione", None),
        ("A8", "Luce libera da terra (m) - promemoria", "B8", 0.8, "Cfg_Luce", None),
        ("A9", "Lungh. fondazione [Agri] (suffisso)", "B9", "30", "Cfg_LenFond", '"' + ",".join(LEN_FOND) + '"'),
        ("A10", "Lungh. pilastro cuscinetto [Basso]", "B10", "30", "Cfg_LenCusc", '"' + ",".join(LEN_CUSC) + '"'),
        ("A11", "Lungh. pilastro motore [Basso]", "B11", "30", "Cfg_LenMot", '"' + ",".join(LEN_MOT) + '"'),
        ("A12", "Pannelli FV nel preventivo", "B12", "No", "Cfg_Pannelli", '"No,Sì"'),
        ("A13", "Margine %", "B13", 0, "Cfg_Margine", None),
    ]
    for acell, label, bcell, val, name, dv in rows_par:
        cfg[acell] = label
        cfg[bcell] = val
        wb.defined_names[name] = DefinedName(name, attr_text=f"CONFIGURATORE!${bcell[0]}${bcell[1:]}")
        if dv:
            d = DataValidation(type="list", formula1=dv, allow_blank=False, showDropDown=False)
            cfg.add_data_validation(d)
            d.add(cfg[bcell])
    # metri accanto alle lunghezze
    cfg["C9"] = '="= "&VALUE(Cfg_LenFond)/10&" m"'
    cfg["C10"] = '="= "&VALUE(Cfg_LenCusc)/10&" m"'
    cfg["C11"] = '="= "&VALUE(Cfg_LenMot)/10&" m"'
    cfg["A15"] = "Trave (frammento codice)"
    cfg["B15"] = '=IF(Cfg_Trave="IPEA 140","IPEA140","IPE140")'
    wb.defined_names["Cfg_TraveFrag"] = DefinedName("Cfg_TraveFrag", attr_text="CONFIGURATORE!$B$15")

    cfg["A18"] = "TIPOLOGIE — inserire il numero di tracker"
    cfg["A18"].font = H
    cfg["A19"] = "Moduli / tracker"; cfg["B19"] = "N. tracker"
    cfg["A19"].font = H; cfg["B19"].font = H
    for i, s in enumerate(SIZES):
        rr = 20 + i
        cfg[f"A{rr}"] = s
        cfg[f"B{rr}"] = 0
        wb.defined_names[f"Cfg_n{s}"] = DefinedName(f"Cfg_n{s}", attr_text=f"CONFIGURATORE!$B${rr}")
    cfg["A26"] = "Totale tracker"; cfg["B26"] = "=SUM(B20:B25)"
    cfg["A26"].font = H

    cfg["A29"] = "RIEPILOGO PREVENTIVO"
    cfg["A29"].font = Font(bold=True, size=12)
    last = 344
    cfg["A30"] = "Materiale (da DISTINTA PROGETTO)"
    cfg["B30"] = f"=SUM('DISTINTA PROGETTO'!$P$2:$P${last})"
    cfg["A31"] = f"Quadro motore (forfait {QUADRO_MOTORE_EUR} €/tracker)"
    cfg["B31"] = f"={QUADRO_MOTORE_EUR}*B26"
    cfg["A32"] = "Costo totale"; cfg["B32"] = "=B30+B31"
    cfg["A33"] = "Margine"; cfg["B33"] = "=B32*Cfg_Margine/100"
    cfg["A34"] = "Prezzo di vendita"; cfg["B34"] = "=B32+B33"
    cfg["A32"].font = H; cfg["B32"].font = H
    cfg["A34"].font = Font(bold=True, size=12); cfg["B34"].font = Font(bold=True, size=12)

    cfg["A37"] = "Materiale per tipologia"
    cfg["A37"].font = H
    for i, s in enumerate(SIZES):
        rr = 38 + i
        col = get_column_letter(8 + i)  # H..M in DISTINTA PROGETTO
        cfg[f"A{rr}"] = f"{s} moduli"
        cfg[f"B{rr}"] = (
            f"=SUMPRODUCT('DISTINTA PROGETTO'!${col}$2:${col}${last},"
            f"'DISTINTA PROGETTO'!$O$2:$O${last})*Cfg_n{s}"
        )

    for cell in ("B30", "B31", "B32", "B33", "B34", *[f"B{38+i}" for i in range(6)]):
        cfg[cell].number_format = '#,##0.00 "€"'
    cfg.column_dimensions["A"].width = 40
    cfg.column_dimensions["B"].width = 16
    cfg.column_dimensions["C"].width = 12

    # ------------------------------------------------------------------ DISTINTA PROGETTO
    dp = wb.create_sheet("DISTINTA PROGETTO", 1)
    headers = ["Assieme", "Sezione", "Cod. distinta", "Descrizione", "Codice progetto",
               "Tag", "Guard", "q12", "q18", "q20", "q22", "q36", "q40",
               "Q.tà progetto", "Prezzo €", "Importo €", "In preventivo"]
    for j, h in enumerate(headers, start=1):
        c = dp.cell(1, j, h); c.font = H
    dp.freeze_panes = "A2"

    AGRI_GUARD = {"FOND_MOT", "FOND_CUSC", "GIUNZ"}
    r_out = 2
    for r in range(2, 345):
        a = dc.cell(r, 1).value
        if a is None and dc.cell(r, 4).value is None and dc.cell(r, 5).value is None:
            continue
        tag, d = tags[r]
        qbs = qty_by_size(dc, r)
        active = tag not in ("SKIP", "EXCL")

        dp.cell(r_out, 1, f"='DISTINTA COMPLETA'!A{r}")
        dp.cell(r_out, 2, f"='DISTINTA COMPLETA'!B{r}")
        dp.cell(r_out, 3, f"='DISTINTA COMPLETA'!D{r}")
        dp.cell(r_out, 4, f"='DISTINTA COMPLETA'!E{r}")

        # --- Codice progetto -------------------------------------------------
        if tag in ("FOND_MOT", "FOND_CUSC"):
            cod = '="TTS.PF.001."&Cfg_TraveFrag&"."&Cfg_LenFond'
        elif tag == "PIL_MOT":
            cod = '="TTS.PM.001."&Cfg_TraveFrag&"."&IF(Cfg_Struttura="Agrivoltaico","22",Cfg_LenMot)'
        elif tag == "PIL_CUSC":
            cod = '="TTS.PC.001."&Cfg_TraveFrag&"."&IF(Cfg_Struttura="Agrivoltaico","23",Cfg_LenCusc)'
        elif tag == "ASSE_PROF":
            cod = f'=IF(Cfg_SpAsse="2,5 mm","{d}.25","{d}")'
        elif tag == "OMEGA_SUP":
            cod = '=IF(Cfg_SpOmega="1,8 mm","TTS.AR.3014.18","TTS.AR.3014")'
        else:
            cod = f"='DISTINTA COMPLETA'!D{r}"
        dp.cell(r_out, 5, cod)
        dp.cell(r_out, 6, tag or "-")

        # --- Guard ---------------------------------------------------------
        if not active:
            guard = "=0"
        elif tag in AGRI_GUARD:
            guard = '=IF(Cfg_Struttura="Agrivoltaico",1,0)'
        elif tag == "PANNELLI":
            guard = '=IF(Cfg_Pannelli="Sì",1,0)'
        else:
            guard = "=1"
        dp.cell(r_out, 7, guard)

        # --- q per taglia (valore x guard) ------------------------------
        for i, s in enumerate(SIZES):
            lit = qbs[s] if active else 0
            lit = int(lit) if float(lit).is_integer() else lit
            dp.cell(r_out, 8 + i, f"=$G{r_out}*{lit}")

        dp.cell(r_out, 14,
                f"=Cfg_n12*H{r_out}+Cfg_n18*I{r_out}+Cfg_n20*J{r_out}"
                f"+Cfg_n22*K{r_out}+Cfg_n36*L{r_out}+Cfg_n40*M{r_out}")
        dp.cell(r_out, 15, f"=IFERROR(VLOOKUP($E{r_out},LISTINO!$B:$G,6,FALSE),0)")
        dp.cell(r_out, 16, f"=$N{r_out}*$O{r_out}")
        dp.cell(r_out, 17, f'=IF($N{r_out}>0,"SI","")')
        dp.cell(r_out, 15).number_format = "#,##0.00"
        dp.cell(r_out, 16).number_format = "#,##0.00"
        r_out += 1

    dp.auto_filter.ref = f"A1:Q{r_out - 1}"
    for col, w in (("A", 22), ("B", 16), ("C", 24), ("D", 44), ("E", 24)):
        dp.column_dimensions[col].width = w
    for col in ("F", "G", "H", "I", "J", "K", "L", "M"):
        dp.column_dimensions[col].hidden = True

    wb.save(out)
    print(f"\nCONFIGURATORE + DISTINTA PROGETTO ({r_out - 2} righe) scritti.")

    # ------------------------------------------------------------------ verifica
    # I costi delle schede sono formule: leggo i valori in cache dal file
    # originale salvato da Excel (i file prodotti da openpyxl non hanno cache).
    pristine = [
        p for p in glob.glob(os.path.join(REPO, "LISTINO componenti tracker TTS 1303 - rev.*.xlsx"))
        if "vecchio" not in p.lower()
        and "collegato alle schede" not in p
        and "configuratore" not in p
    ]
    if pristine:
        verify(max(pristine, key=lambda p: int(re.search(r"rev\.(\d+)", os.path.basename(p)).group(1))), tags)
    else:
        print("(verifica saltata: nessun file originale salvato da Excel per i costi delle schede)")


def verify(src, tags):
    """Ricalcola in Python il totale per una configurazione di esempio."""
    print(f"\n(verifica prezzi da: {os.path.basename(src)})")
    wb = load_workbook(src, data_only=True)
    dc = wb["DISTINTA COMPLETA"]

    # mappa prezzi = LISTINO!G calcolato dalle schede
    def elenco(name, hdr, code_c, cost_cs):
        ws = wb[name]; o = {}
        for r in range(hdr + 1, ws.max_row + 1):
            k = ws.cell(r, code_c).value
            if k in (None, ""):
                continue
            t = 0.0; ok = False
            for c in cost_cs:
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    t += v; ok = True
            o[str(k)] = t if ok else None
        return o
    st = elenco("SCHEDA STRUTTURA", 16, 1, [6, 7])
    pz = elenco("SCHEDA PEZZI SPECIALI", 37, 1, [5, 6])
    el = elenco("SCHEDA ELETTROMECCANICA", 4, 1, [4])
    bu = elenco("SCHEDA BULLONERIA", 5, 1, [4])

    def price(code):
        for s in (st, pz, el, bu):
            if code in s and s[code] is not None:
                return s[code]
        return 0.0

    counts = {12: 4, 18: 1, 20: 20, 22: 5, 36: 34, 40: 5}
    struttura, trave, sp_asse, sp_omega, len_fond, pannelli = "AGRI", "IPE140", "3,0", "2,0", "30", False
    tot = 0.0
    for r in range(2, 345):
        tag, d = tags.get(r, ("", ""))
        if tag in ("SKIP", "EXCL"):
            continue
        if tag in ("FOND_MOT", "FOND_CUSC", "GIUNZ") and struttura != "AGRI":
            continue
        if tag == "PANNELLI" and not pannelli:
            continue
        # codice progetto
        if tag in ("FOND_MOT", "FOND_CUSC"):
            cod = f"TTS.PF.001.{trave}.{len_fond}"
        elif tag == "PIL_MOT":
            cod = f"TTS.PM.001.{trave}.22"
        elif tag == "PIL_CUSC":
            cod = f"TTS.PC.001.{trave}.23"
        elif tag == "ASSE_PROF":
            cod = d + (".25" if sp_asse == "2,5" else "")
        elif tag == "OMEGA_SUP":
            cod = "TTS.AR.3014.18" if sp_omega == "1,8" else "TTS.AR.3014"
        else:
            cod = d
        pr = price(str(cod)) if cod else 0.0
        a = (dc.cell(r, 1).value or "").strip()
        b = (dc.cell(r, 2).value or "")
        base = dc.cell(r, 7).value
        base = base if isinstance(base, (int, float)) else 0
        for s, n in counts.items():
            v = dc.cell(r, CFG_COL[s]).value
            q = v if isinstance(v, (int, float)) else base
            if a == "ASSI DI ROTAZIONE 1303" and b != f"{s} MODULI 1303":
                q = 0
            tot += n * q * pr
    quadro = QUADRO_MOTORE_EUR * sum(counts.values())
    print("\n--- verifica (Agri, IPE140, 3,0 mm, omega 2,0, fondazione .30, no pannelli) ---")
    print(f"  tracker ......... {sum(counts.values())}")
    print(f"  materiale ....... {tot:,.2f} €")
    print(f"  quadro motore ... {quadro:,.2f} €")
    print(f"  COSTO TOTALE .... {tot + quadro:,.2f} €")


if __name__ == "__main__":
    main()
